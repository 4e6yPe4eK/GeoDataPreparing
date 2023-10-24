import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QPushButton, QGridLayout, QLineEdit,
                             QFileDialog, QHBoxLayout, QRadioButton, QButtonGroup)
from PyQt5.QtCore import QThread, QObject, pyqtSignal
import openpyxl
from widgets.checkboxlistwidget import CheckboxListWidget
from functools import partial
import logging

from sentinel import communicator, const


logging.basicConfig(filename="log.log", level=logging.INFO)


class Worker(QObject):
    finished = pyqtSignal()
    progressChanged = pyqtSignal(int)

    def run(self, data):
        self.progressChanged.emit(0)
        communicator.parse_directories(data, self.progressChanged.emit)
        self.finished.emit()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Geo")
        self.resize(800, 600)
        self.message("", 1)
        self.new_thread = None

        self.widget = QWidget(self)
        self.setCentralWidget(self.widget)

        self.layout = QGridLayout(self.widget)
        self.widget.setLayout(self.layout)

        self.directory_line = QLineEdit(self.widget)
        self.directory_line.setPlaceholderText("Путь к папке c исходными данными")
        self.layout.addWidget(self.directory_line, 0, 0, 1, 1)

        self.directory_button = QPushButton("Выбрать", self.widget)
        self.directory_button.clicked.connect(self.directory_button_clicked)
        self.layout.addWidget(self.directory_button, 0, 1, 1, 1)

        self.r_button_layout = QHBoxLayout()
        self.layout.addLayout(self.r_button_layout, 1, 0, 1, 2)

        self.r10_button = QRadioButton("R10m", self.widget)
        self.r10_button.setChecked(True)
        self.r_button_layout.addWidget(self.r10_button)

        self.r20_button = QRadioButton("R20m", self.widget)
        self.r_button_layout.addWidget(self.r20_button)

        self.r60_button = QRadioButton("R60m", self.widget)
        self.r_button_layout.addWidget(self.r60_button)

        self.r_button_group = QButtonGroup()
        self.r_button_group.addButton(self.r10_button)
        self.r_button_group.addButton(self.r20_button)
        self.r_button_group.addButton(self.r60_button)
        self.r_button_group.buttonClicked.connect(self.r_button_group_clicked)

        self.shape_line = QLineEdit(self.widget)
        self.shape_line.setPlaceholderText("Путь к shape-файлу")
        self.layout.addWidget(self.shape_line, 2, 0, 1, 1)

        self.shape_button = QPushButton("Выбрать", self.widget)
        self.shape_button.clicked.connect(self.shape_button_clicked)
        self.layout.addWidget(self.shape_button, 2, 1, 1, 1)

        self.output_line = QLineEdit(self.widget)
        self.output_line.setPlaceholderText("Путь к папке для сохранения")
        self.layout.addWidget(self.output_line, 3, 0, 1, 1)

        self.output_button = QPushButton("Выбрать", self.widget)
        self.output_button.clicked.connect(self.output_button_clicked)
        self.layout.addWidget(self.output_button, 3, 1, 1, 1)

        self.match_line = QLineEdit(self.widget)
        self.match_line.setPlaceholderText("Путь к файлу для сопоставления названий полей")
        self.layout.addWidget(self.match_line, 4, 0, 1, 1)

        self.match_button = QPushButton("Выбрать", self.widget)
        self.match_button.clicked.connect(self.match_button_clicked)
        self.layout.addWidget(self.match_button, 4, 1, 1, 1)

        self.match_hash = 0
        self.match_data = {}

        self.choice_button_layout = QHBoxLayout()
        self.layout.addLayout(self.choice_button_layout, 5, 0, 1, 2)

        self.field_choice_button = QPushButton("Поля", self.widget)
        self.field_choice_button.clicked.connect(self.field_choice_button_clicked)
        self.choice_button_layout.addWidget(self.field_choice_button)
        self.field_choice_widget = CheckboxListWidget(self.widget)

        self.coefficient_choice_button = QPushButton("Коэффициенты", self.widget)
        self.coefficient_choice_button.clicked.connect(self.coefficient_choice_button_clicked)
        self.choice_button_layout.addWidget(self.coefficient_choice_button)
        self.coefficient_choice_widget = CheckboxListWidget(self.widget)
        self.r_button_group_clicked()

        self.start_button = QPushButton("Начать", self.widget)
        self.start_button.clicked.connect(self.start_button_clicked)
        self.layout.addWidget(self.start_button, 6, 0, 1, 2)

    def load_match_data(self):
        match_path = self.match_line.text()
        self.match_hash = hash(match_path)
        self.match_data = {}
        wb: openpyxl.workbook.Workbook = openpyxl.load_workbook(match_path, read_only=True)
        sheet = wb.active
        for row in range(1, sheet.max_row + 1):
            self.match_data[int(sheet.cell(row=row, column=1).value)] = str(sheet.cell(row=row, column=2).value)

    def message(self, text, time=0):
        self.statusBar().showMessage(text, time)

    def directory_button_clicked(self):
        directory = QFileDialog.getExistingDirectory(self, "Выбрать папку")
        self.directory_line.setText(directory)

    def output_button_clicked(self):
        directory = QFileDialog.getExistingDirectory(self, "Выбрать папку")
        self.output_line.setText(directory)

    def shape_button_clicked(self):
        directory = QFileDialog.getOpenFileName(self, "Выбрать shape-файл", filter="Shape-файл (*.shp)")
        self.shape_line.setText(directory[0])

    def field_choice_button_clicked(self):
        match_path = self.match_line.text()
        if not match_path:
            self.message("Ошибка: файл для сопоставления не выбран", 3000)
            return
        if self.match_hash != hash(match_path):
            self.load_match_data()
            self.field_choice_widget.set_choices(sorted(set(self.match_data.values())))
        self.field_choice_widget.exec()

    def coefficient_choice_button_clicked(self):
        self.coefficient_choice_widget.exec()

    def match_button_clicked(self):
        directory = QFileDialog.getOpenFileName(self, "Выбрать Excel-файл", filter="Excel-файл (*.xls, *.xlsx)")
        self.match_line.setText(directory[0])

    def start_button_clicked(self):
        directory = self.directory_line.text()
        shape = self.shape_line.text()
        output = self.output_line.text()
        resolution = self.r_button_group.checkedButton().text()
        fields = self.field_choice_widget.selected_item_texts()
        coefficients = self.coefficient_choice_widget.selected_item_texts()
        match = self.match_line.text()
        if directory == "":
            self.message("Ошибка: папка с исходными данными не выбрана", 3000)
        elif shape == "":
            self.message("Ошибка: shape-файл не выбран", 3000)
        elif output == "":
            self.message("Ошибка: папка для сохранения не выбрана", 3000)
        elif match == "":
            pass
        else:
            if self.match_hash != hash(match):
                self.load_match_data()
                self.field_choice_widget.set_choices(sorted(set(self.match_data.values())))
                fields = self.field_choice_widget.selected_item_texts()
            data = {
                "directory": directory,
                "shape": shape,
                "output": output,
                "resolution": resolution,
                "fields": fields,
                "coefficients": coefficients,
                "match_fields": self.match_data,
            }
            self.new_thread = QThread()
            worker = Worker()
            worker.moveToThread(self.new_thread)
            self.new_thread.started.connect(partial(worker.run, data))
            worker.finished.connect(self.new_thread.quit)
            worker.finished.connect(lambda: self.start_button.setText("Начать"))
            worker.finished.connect(lambda: self.start_button.setEnabled(True))
            worker.finished.connect(lambda: self.message("Обработка завершена", 3000))
            worker.progressChanged.connect(self.progress_changed)
            self.new_thread.finished.connect(worker.deleteLater)
            self.new_thread.finished.connect(self.new_thread.deleteLater)
            self.new_thread.start()
            self.start_button.setText("Идет обработка...")
            self.start_button.setEnabled(False)

    def r_button_group_clicked(self):
        btn = self.r_button_group.checkedButton()
        choices = []
        if btn is self.r10_button:
            choices = const.coefficient_names_r10
        if btn is self.r20_button:
            choices = const.coefficient_names_r20
        if btn is self.r60_button:
            choices = const.coefficient_names_r60
        self.coefficient_choice_widget.set_choices(choices)

    def progress_changed(self, percent):
        self.message(f"Завершено на {percent}%")
